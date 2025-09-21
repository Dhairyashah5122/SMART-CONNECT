"""
Data Extraction and Export Engine
MIT License - Westcliff University Property
"""

from typing import Dict, List, Optional, Any, Union
from datetime import datetime, date
from enum import Enum
import json
import csv
import io
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from pydantic import BaseModel, Field
import base64

from .data_mining import DataMiningEngine, SearchQuery, SearchResult


class ExportFormat(str, Enum):
    JSON = "json"
    CSV = "csv"
    EXCEL = "excel"
    PDF = "pdf"
    XML = "xml"


class ExportOptions(BaseModel):
    format: ExportFormat
    include_headers: bool = True
    include_metadata: bool = True
    include_relations: bool = False
    flatten_json: bool = True
    custom_filename: Optional[str] = None
    template_name: Optional[str] = None
    compression: bool = False


class ExportResult(BaseModel):
    filename: str
    format: ExportFormat
    size_bytes: int
    record_count: int
    file_data: str  # Base64 encoded for binary formats
    download_url: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.now)


class DataExtractionEngine:
    """Advanced data extraction and export engine"""
    
    def __init__(self, mining_engine: DataMiningEngine):
        self.mining_engine = mining_engine
        self.templates = self._initialize_templates()
    
    def _initialize_templates(self) -> Dict[str, Dict[str, Any]]:
        """Initialize export templates"""
        return {
            "student_report": {
                "fields": [
                    "student_id", "first_name", "last_name", "email", 
                    "program", "specialization", "gpa", "status",
                    "enrollment_date", "expected_graduation_date"
                ],
                "title": "Student Report",
                "description": "Comprehensive student information"
            },
            "mentor_directory": {
                "fields": [
                    "first_name", "last_name", "email", "company_name", 
                    "job_title", "department", "industry", "years_of_experience"
                ],
                "title": "Mentor Directory",
                "description": "Professional mentor contact directory"
            },
            "project_catalog": {
                "fields": [
                    "title", "description", "project_type", "difficulty_level",
                    "status", "start_date", "end_date", "max_students"
                ],
                "title": "Project Catalog",
                "description": "Available capstone projects"
            },
            "analytics_summary": {
                "fields": [
                    "entity_type", "total_count", "active_count", 
                    "last_updated", "completion_rate"
                ],
                "title": "Analytics Summary",
                "description": "Key metrics and analytics"
            }
        }
    
    async def extract_data(self, query: SearchQuery, 
                          export_options: ExportOptions) -> ExportResult:
        """Extract data based on search query and export in specified format"""
        
        # Execute search query
        search_result = await self.mining_engine.search(query)
        
        # Apply template if specified
        if export_options.template_name and export_options.template_name in self.templates:
            search_result.data = self._apply_template(
                search_result.data, export_options.template_name
            )
        
        # Flatten JSON if requested
        if export_options.flatten_json and export_options.format != ExportFormat.JSON:
            search_result.data = self._flatten_data(search_result.data)
        
        # Generate filename
        filename = self._generate_filename(query, export_options)
        
        # Export based on format
        if export_options.format == ExportFormat.JSON:
            file_data, size = self._export_json(search_result, export_options)
        elif export_options.format == ExportFormat.CSV:
            file_data, size = self._export_csv(search_result, export_options)
        elif export_options.format == ExportFormat.EXCEL:
            file_data, size = self._export_excel(search_result, export_options)
        elif export_options.format == ExportFormat.PDF:
            file_data, size = self._export_pdf(search_result, export_options)
        elif export_options.format == ExportFormat.XML:
            file_data, size = self._export_xml(search_result, export_options)
        else:
            raise ValueError(f"Unsupported export format: {export_options.format}")
        
        # Apply compression if requested
        if export_options.compression:
            file_data, size = self._compress_data(file_data, filename)
            filename += ".zip"
        
        return ExportResult(
            filename=filename,
            format=export_options.format,
            size_bytes=size,
            record_count=len(search_result.data),
            file_data=base64.b64encode(file_data).decode('utf-8') if isinstance(file_data, bytes) else file_data
        )
    
    def _apply_template(self, data: List[Dict[str, Any]], 
                       template_name: str) -> List[Dict[str, Any]]:
        """Apply predefined template to filter fields"""
        template = self.templates[template_name]
        fields = template["fields"]
        
        filtered_data = []
        for record in data:
            filtered_record = {}
            for field in fields:
                if field in record:
                    filtered_record[field] = record[field]
                else:
                    # Handle nested fields (e.g., "user.email")
                    if "." in field:
                        parts = field.split(".")
                        value = record
                        for part in parts:
                            if isinstance(value, dict) and part in value:
                                value = value[part]
                            else:
                                value = None
                                break
                        filtered_record[field] = value
                    else:
                        filtered_record[field] = None
            filtered_data.append(filtered_record)
        
        return filtered_data
    
    def _flatten_data(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Flatten nested JSON structures"""
        flattened_data = []
        
        for record in data:
            flattened_record = {}
            self._flatten_dict(record, flattened_record)
            flattened_data.append(flattened_record)
        
        return flattened_data
    
    def _flatten_dict(self, nested_dict: Dict[str, Any], 
                     flattened_dict: Dict[str, Any], prefix: str = ""):
        """Recursively flatten a nested dictionary"""
        for key, value in nested_dict.items():
            new_key = f"{prefix}{key}" if prefix else key
            
            if isinstance(value, dict):
                self._flatten_dict(value, flattened_dict, f"{new_key}.")
            elif isinstance(value, list) and value and isinstance(value[0], dict):
                # Handle list of dictionaries
                for i, item in enumerate(value):
                    if isinstance(item, dict):
                        self._flatten_dict(item, flattened_dict, f"{new_key}.{i}.")
                    else:
                        flattened_dict[f"{new_key}.{i}"] = item
            elif isinstance(value, list):
                # Handle simple lists
                flattened_dict[new_key] = json.dumps(value) if value else ""
            else:
                flattened_dict[new_key] = value
    
    def _generate_filename(self, query: SearchQuery, 
                          export_options: ExportOptions) -> str:
        """Generate appropriate filename"""
        if export_options.custom_filename:
            return export_options.custom_filename
        
        # Use template name if available
        if export_options.template_name:
            base_name = export_options.template_name
        else:
            base_name = f"{query.entity}_export"
        
        # Add timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        extension = export_options.format.value
        
        return f"{base_name}_{timestamp}.{extension}"
    
    def _export_json(self, result: SearchResult, 
                    options: ExportOptions) -> tuple[str, int]:
        """Export data as JSON"""
        export_data = {
            "data": result.data,
            "metadata": {
                "total_count": result.total_count,
                "page": result.page,
                "page_size": result.page_size,
                "total_pages": result.total_pages,
                "execution_time_ms": result.execution_time_ms,
                "query_info": result.query_info,
                "exported_at": datetime.now().isoformat()
            } if options.include_metadata else None,
            "aggregations": result.aggregations if result.aggregations else None
        }
        
        # Remove None values
        export_data = {k: v for k, v in export_data.items() if v is not None}
        
        json_str = json.dumps(export_data, indent=2, default=str)
        return json_str, len(json_str.encode('utf-8'))
    
    def _export_csv(self, result: SearchResult, 
                   options: ExportOptions) -> tuple[bytes, int]:
        """Export data as CSV"""
        if not result.data:
            return b"", 0
        
        output = io.StringIO()
        
        # Get all possible fieldnames
        fieldnames = set()
        for record in result.data:
            fieldnames.update(record.keys())
        fieldnames = sorted(list(fieldnames))
        
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        
        if options.include_headers:
            writer.writeheader()
        
        # Write data
        for record in result.data:
            # Convert complex types to strings
            csv_record = {}
            for key, value in record.items():
                if isinstance(value, (dict, list)):
                    csv_record[key] = json.dumps(value, default=str)
                elif isinstance(value, (date, datetime)):
                    csv_record[key] = value.isoformat()
                else:
                    csv_record[key] = value
            writer.writerow(csv_record)
        
        # Add metadata if requested
        if options.include_metadata:
            output.write("\n# Metadata\n")
            output.write(f"# Total Count: {result.total_count}\n")
            output.write(f"# Execution Time: {result.execution_time_ms}ms\n")
            output.write(f"# Exported At: {datetime.now().isoformat()}\n")
        
        csv_content = output.getvalue().encode('utf-8')
        return csv_content, len(csv_content)
    
    def _export_excel(self, result: SearchResult, 
                     options: ExportOptions) -> tuple[bytes, int]:
        """Export data as Excel"""
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Data Export"
        
        if not result.data:
            return b"", 0
        
        # Get all fieldnames
        fieldnames = set()
        for record in result.data:
            fieldnames.update(record.keys())
        fieldnames = sorted(list(fieldnames))
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Write headers
        if options.include_headers:
            for col, fieldname in enumerate(fieldnames, 1):
                cell = worksheet.cell(row=1, column=col, value=fieldname)
                cell.font = header_font
                cell.fill = header_fill
            row_offset = 2
        else:
            row_offset = 1
        
        # Write data
        for row_idx, record in enumerate(result.data, row_offset):
            for col_idx, fieldname in enumerate(fieldnames, 1):
                value = record.get(fieldname)
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, default=str)
                elif isinstance(value, (date, datetime)):
                    value = value.isoformat()
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Add metadata sheet if requested
        if options.include_metadata:
            metadata_sheet = workbook.create_sheet("Metadata")
            metadata_sheet.cell(row=1, column=1, value="Total Count")
            metadata_sheet.cell(row=1, column=2, value=result.total_count)
            metadata_sheet.cell(row=2, column=1, value="Execution Time (ms)")
            metadata_sheet.cell(row=2, column=2, value=result.execution_time_ms)
            metadata_sheet.cell(row=3, column=1, value="Exported At")
            metadata_sheet.cell(row=3, column=2, value=datetime.now().isoformat())
        
        # Save to bytes
        output = io.BytesIO()
        workbook.save(output)
        excel_content = output.getvalue()
        return excel_content, len(excel_content)
    
    def _export_pdf(self, result: SearchResult, 
                   options: ExportOptions) -> tuple[bytes, int]:
        """Export data as PDF"""
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        story = []
        
        # Title
        title = Paragraph("Data Export Report", title_style)
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Metadata
        if options.include_metadata:
            metadata_data = [
                ["Total Records", str(result.total_count)],
                ["Execution Time", f"{result.execution_time_ms}ms"],
                ["Exported At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
            ]
            
            metadata_table = Table(metadata_data)
            metadata_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ]))
            story.append(metadata_table)
            story.append(Spacer(1, 12))
        
        # Data table
        if result.data:
            # Limit fields for PDF readability
            fieldnames = list(result.data[0].keys())[:6]  # First 6 fields
            
            # Prepare table data
            table_data = []
            if options.include_headers:
                table_data.append(fieldnames)
            
            for record in result.data[:50]:  # Limit to 50 records for PDF
                row = []
                for fieldname in fieldnames:
                    value = record.get(fieldname, "")
                    if isinstance(value, (dict, list)):
                        value = str(value)[:50] + "..." if len(str(value)) > 50 else str(value)
                    elif isinstance(value, (date, datetime)):
                        value = value.strftime("%Y-%m-%d")
                    row.append(str(value)[:30])  # Truncate for readability
                table_data.append(row)
            
            # Create table
            data_table = Table(table_data)
            data_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(data_table)
        
        # Build PDF
        doc.build(story)
        pdf_content = output.getvalue()
        return pdf_content, len(pdf_content)
    
    def _export_xml(self, result: SearchResult, 
                   options: ExportOptions) -> tuple[str, int]:
        """Export data as XML"""
        xml_lines = ['<?xml version="1.0" encoding="UTF-8"?>']
        xml_lines.append('<export>')
        
        # Metadata
        if options.include_metadata:
            xml_lines.append('  <metadata>')
            xml_lines.append(f'    <total_count>{result.total_count}</total_count>')
            xml_lines.append(f'    <execution_time_ms>{result.execution_time_ms}</execution_time_ms>')
            xml_lines.append(f'    <exported_at>{datetime.now().isoformat()}</exported_at>')
            xml_lines.append('  </metadata>')
        
        # Data
        xml_lines.append('  <data>')
        for record in result.data:
            xml_lines.append('    <record>')
            for key, value in record.items():
                # Sanitize key for XML
                clean_key = re.sub(r'[^a-zA-Z0-9_]', '_', str(key))
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, default=str)
                elif isinstance(value, (date, datetime)):
                    value = value.isoformat()
                # Escape XML special characters
                value = str(value).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                xml_lines.append(f'      <{clean_key}>{value}</{clean_key}>')
            xml_lines.append('    </record>')
        xml_lines.append('  </data>')
        xml_lines.append('</export>')
        
        xml_content = '\n'.join(xml_lines)
        return xml_content, len(xml_content.encode('utf-8'))
    
    def _compress_data(self, data: Union[str, bytes], filename: str) -> tuple[bytes, int]:
        """Compress data using ZIP"""
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            if isinstance(data, str):
                zip_file.writestr(filename, data.encode('utf-8'))
            else:
                zip_file.writestr(filename, data)
        
        compressed_data = zip_buffer.getvalue()
        return compressed_data, len(compressed_data)
    
    def get_available_templates(self) -> Dict[str, Dict[str, Any]]:
        """Get all available export templates"""
        return self.templates